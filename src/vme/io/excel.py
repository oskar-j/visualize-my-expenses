"""Excel workbooks (.xlsx) -- needs the optional ``openpyxl`` extra."""

from __future__ import annotations

from typing import List, Optional

from ..models import Expense
from ..tools import coerce_row
from .base import Format, LoaderError, register, require

__all__ = ["load_excel"]


def load_excel(path: str, sheet: Optional[str] = None, default_currency: str = "USD",
               dayfirst: Optional[bool] = None, **_: object) -> List[Expense]:
    require("openpyxl", "excel", "Excel")
    from openpyxl import load_workbook

    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise LoaderError(f"cannot open {path}: {exc}") from exc

    try:
        worksheet = book[sheet] if sheet else book[book.sheetnames[0]]
    except KeyError as exc:
        raise LoaderError(
            f"{path} has no sheet {sheet!r}; available: {', '.join(book.sheetnames)}"
        ) from exc

    rows = [row for row in worksheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)]
    if not rows:
        raise LoaderError(f"{path}: sheet {worksheet.title!r} is empty")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    expenses: List[Expense] = []
    for number, row in enumerate(rows[1:], start=2):
        record = {key: value for key, value in zip(header, row) if key}
        if not any(value not in (None, "") for value in record.values()):
            continue
        try:
            expenses.append(coerce_row(record, default_currency, dayfirst))
        except (ValueError, TypeError) as exc:
            raise LoaderError(f"{path}:{worksheet.title}:{number}: {exc}") from exc

    if not expenses:
        raise LoaderError(f"{path}: sheet {worksheet.title!r} has a header but no rows")
    return expenses


register(Format("excel", "Excel workbook (.xlsx), first sheet unless --sheet is given",
                ("xlsx", "xlsm"), load_excel, requires="openpyxl"))
